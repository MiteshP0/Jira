import base64
import time
from openpyxl import Workbook
import jira.client
from jira.client import JIRA
import datetime
from datetime import date




# -- Get todays date
todayDate = datetime.datetime.now()
todayDate_count = date(int(todayDate.year),int(todayDate.month),int(todayDate.day))

#-- URL to the Query : https://cibu-jira:8443/rest/api/2/search?jql=category=%22Spark%20Call%20Projects%22%20AND%20created%3E=-360d

#-- Create workbook to write excel
wb = Workbook()
ws = wb.active

options = {'server': 'https://cibu-jira.cisco.com:8443'}
#-- authorization to access jira, enter your password in password field and username in username field.
password = base64.b64encode("password")
jira = JIRA(options, basic_auth=('username', base64.b64decode(password)))
# jira = JIRA(options)


#-- Generates list of issues in the excel file
# ---generateIssues(apiurl, startnum):
# --- @Param: apiurl - API Query you wish to use
# --- @Param: startnum - Number from which you want to start writing from (ie if there are 2000 results in the query, and you want to start at 1000th issue then pass in 1000)
def generateIssues(apiurl, startNum) :
    issues = jira.search_issues(apiurl, startAt=startNum, maxResults=1000)
    for k, i in enumerate(issues):

        # resolDate = ''
        # createdDate = ''
        # -- Set temporary date which will get over written. Only used to make the variable global in the entire loop and to set it as a date object.
        createdDate_count = date(1996,02,06)
        resolDate_count = date(1996,02,06)
        issue = jira.issue(i.key)
        k = k + (startNum+3)
        column_row = 'A' + str(k)
        ws[column_row] = str(issue)
        if 'priority' in issue.raw['fields'] and issue.raw['fields']['priority'] is not None:
            if 'name' in issue.raw['fields']['priority']:
                priority = str(issue.raw['fields']['priority']['name']).strip()
                column_row = 'B' + str(k)
                ws[column_row] = priority
            else:
                priority = 'unknown'
                column_row = 'B' + str(k)
                ws[column_row] = priority
        else:
            priority = 'unknown'
            column_row = 'B' + str(k)
            ws[column_row] = priority
        if 'issuetype' in issue.raw['fields'] and issue.raw['fields']['issuetype'] is not None:
            if 'name' in issue.raw['fields']['issuetype']:
                issuetype = str(issue.raw['fields']['issuetype']['name']).strip()
                column_row = 'C' + str(k)
                ws[column_row] = issuetype
            else:
                issuetype = 'unknown'
                column_row = 'C' + str(k)
                ws[column_row] = issuetype
        else:
            issuetype = 'unknown'
            column_row = 'C' + str(k)
            ws[column_row] = issuetype
        if 'status' in issue.raw['fields'] and issue.raw['fields']['status'] is not None:
            if 'name' in issue.raw['fields']['status']:
                status = str(issue.raw['fields']['status']['name']).strip()
                column_row = 'D' + str(k)
                ws[column_row] = status
            else:
                status = 'unknown'
                column_row = 'D' + str(k)
                ws[column_row] = status
        else:
            status = 'unknown'
            column_row = 'D' + str(k)
            ws[column_row] = status
        if 'assignee' in issue.raw['fields'] and issue.raw['fields']['assignee'] is not None:
            if 'emailAddress' in issue.raw['fields']['assignee']:
                assignee = str(issue.raw['fields']['assignee']['emailAddress']).strip()
                column_row = 'E' + str(k)
                ws[column_row] = assignee
            else:
                assignee = 'unknown'
                column_row = 'E' + str(k)
                ws[column_row] = assignee
        else:
            assignee = 'unknown'
            column_row = 'E' + str(k)
            ws[column_row] = assignee
        if 'reporter' in issue.raw['fields'] and issue.raw['fields']['reporter'] is not None:
            if 'emailAddress' in issue.raw['fields']['reporter']:
                reporter = str(issue.raw['fields']['reporter']['emailAddress']).strip()
                column_row = 'F' + str(k)
                ws[column_row] = reporter
            else:
                reporter = 'unknown'
                column_row = 'F' + str(k)
                ws[column_row] = reporter
        else:
            reporter = 'unknown'
            column_row = 'F' + str(k)
            ws[column_row] = reporter
        if 'customfield_11711' in issue.raw['fields'] and issue.raw['fields']['customfield_11711'] is not None:
            if 'value' in issue.raw['fields']['customfield_11711']:
                custAffect = str(issue.raw['fields']['customfield_11711']['value']).strip()
                column_row = 'G' + str(k)
                ws[column_row] = custAffect
            else:
                custAffect = 'unknown'
                column_row = 'G' + str(k)
                ws[column_row] = custAffect
        else:
            custAffect = 'unknown'
            column_row = 'G' + str(k)
            ws[column_row] = custAffect
        if 'customfield_10641' in issue.raw['fields'] and issue.raw['fields']['customfield_10641'] is not None:
            if 'value' in issue.raw['fields']['customfield_10641']:
                origination = str(issue.raw['fields']['customfield_10641']['value']).strip()
                column_row = 'H' + str(k)
                ws[column_row] = origination
            else:
                origination = 'unknown'
                column_row = 'H' + str(k)
                ws[column_row] = origination
        else:
            origination = 'unknown'
            column_row = 'H' + str(k)
            ws[column_row] = origination
        if 'labels' in issue.raw['fields'] and issue.raw['fields']['labels'] is not None:
            label = str(issue.raw['fields']['labels']).strip()
            column_row = 'I' + str(k)
            ws[column_row] = label
        else:
            label = 'unknown'
            column_row = 'I' + str(k)
            ws[column_row] = label
        if 'created' in issue.raw['fields'] and issue.raw['fields']['created'] is not None:
            createdDate = str(issue.raw['fields']['created']).strip()
            createdDateFormat = createdDate[0:10]
            # -- date(year,month,day)
            createdDate_count = date(int(createdDateFormat[0:4]),int(createdDateFormat[5:7]),
                                     int(createdDateFormat[8:]))
            column_row = 'J' + str(k)
            ws[column_row] = createdDateFormat
        else:
            createdDate = 'unknown'
            # createdDate_count = 0
            column_row = 'J' + str(k)
            ws[column_row] = createdDate
        if 'resolutiondate' in issue.raw['fields'] and issue.raw['fields']['resolutiondate'] is not None:
            resolDate = str(issue.raw['fields']['resolutiondate']).strip()
            resolDateFormat = resolDate[0:10]
            # -- date(year,month,day)
            resolDate_count = date(int(resolDateFormat[0:4]), int(resolDateFormat[5:7]),
                                     int(resolDateFormat[8:]))
            column_row = 'K' + str(k)
            ws[column_row] = resolDateFormat
        else:
            resolDate = 'unknown'
            # resolDate_count = 0
            column_row = 'K' + str(k)
            ws[column_row] = resolDate
        if resolDate != 'unknown' :
            column_row = 'L' + str(k)
            deltadate = resolDate_count-createdDate_count
            ws[column_row] = str(deltadate.days)
        if createdDate != 'unknown':
            column_row = 'M' + str(k)
            deltadate = todayDate_count - createdDate_count
            ws[column_row] = str(deltadate.days)

def main() :
    contentCategory = ['Key', 'Priority', 'Issue Type', 'Status', 'Assignee', 'Reporter', 'Customer Affecting?',
                       'Origination', 'Labels','Created', 'Resolution Date', 'MTTR (days)' , 'Average Open (days)']
    columnCell = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1','M1']
    #-- Assign columns to the excel file
    for i, k in zip(contentCategory, columnCell):
        ws[k] = i
    url = "( (category=\"Spark Call Projects\" AND created>=-360d) )"
    #-- Generate issues up to 11000+
    generateIssues(url,0)
    # generateIssues(url, 1000)
    # generateIssues(url, 2000)
    # generateIssues(url, 3000)
    # generateIssues(url, 4000)
    # generateIssues(url, 5000)
    # generateIssues(url, 6000)
    # generateIssues(url, 7000)
    # generateIssues(url, 8000)
    # generateIssues(url, 9000)
    # generateIssues(url, 10000)
    # generateIssues(url, 11000)

if __name__ == '__main__':
    start_time = time.time()
    main()
    print "Program took ", (time.time() - start_time) / 60, " minutes to run."


# Save the file
wb.save("jiraspark10.xlsx")
